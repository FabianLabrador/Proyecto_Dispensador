from openpyxl import load_workbook
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook

def guardar_direccionamientos_en_excel(data, ruta_archivo):
    libro = load_workbook(ruta_archivo)
    hoja = libro.active
    hoja.title = "Direccionamientos"

    encabezados = list(data[0].keys())
    hoja.append(encabezados)

    for fila in data:
        hoja.append(list(fila.values()))

    libro.save(ruta_archivo)    

def cargar_datos_excel(ruta_archivo):
    """Carga un archivo Excel como DataFrame."""
    try:
        return pd.read_excel(ruta_archivo)
    except Exception as e:
        raise IOError(f"No se pudo leer el archivo Excel: {e}")

def guardar_respuestas_en_excel(df, respuestas, ruta_archivo, nombre_hoja):
    """
    """
    if len(df) != len(respuestas):
        raise ValueError(f"Cantidad de respuestas ({len(respuestas)}) no coincide con filas del Excel ({len(df)})")

    df['Respuesta_API'] = respuestas

    try:
        df.to_excel(ruta_archivo, sheet_name=nombre_hoja, index=False)
        print(f"[✔] Archivo guardado en hoja '{nombre_hoja}': {ruta_archivo}")
    except Exception as e:
        print(f"[✘] Error al guardar archivo Excel: {e}")
        raise

def cargar_datos_programacion(ruta_archivo):
    """Carga la hoja 'Programacion' desde el archivo Excel."""
    return pd.read_excel(ruta_archivo, sheet_name='Programacion')

def cargar_datos_entrega(ruta_archivo):
    """Carga la hoja 'Entrega' desde el archivo Excel."""
    return pd.read_excel(ruta_archivo, sheet_name='Entrega')

def cargar_datos_reporte_entrega(ruta_archivo):
    return pd.read_excel(ruta_archivo, sheet_name='ReporteEntrega')

def cargar_datos_reporte_facturacion(ruta_archivo):
    return pd.read_excel(ruta_archivo, sheet_name='ReporteFacturacion')

def guardar_reporte_direccionamientos_por_prescripcion(resultados, ruta_archivo, hoja_nombre, encabezados):
    libro = load_workbook(ruta_archivo)
    if hoja_nombre in libro.sheetnames:
        del libro[hoja_nombre]  # Eliminar si ya existe

    hoja = libro.create_sheet(title=hoja_nombre)

    # Encabezados
    for i, encabezado in enumerate(encabezados, 1):
        hoja.cell(row=1, column=i, value=encabezado)

    # Contenido
    fila = 2
    for registro in resultados:
        for record in registro:
            for col, key in enumerate(encabezados, 1):
                hoja.cell(row=fila, column=col, value=str(record.get(key, '')))
            fila += 1

    libro.save(ruta_archivo)

def guardar_reporte_combinado_en_excel(ruta_archivo, encabezados, data_dict, hoja_nombre="ReporteCombinado"):
    """
    Guarda en Excel la información combinada de varios endpoints del Reporteador.
    Crea una hoja nueva con los encabezados prefijados y los datos alineados por ID.
    """
    try:
        # Abrir el archivo existente
        libro = load_workbook(ruta_archivo)

        # Si la hoja ya existe, la eliminamos para regenerarla limpia
        if hoja_nombre in libro.sheetnames:
            std = libro[hoja_nombre]
            libro.remove(std)

        # Crear nueva hoja
        hoja = libro.create_sheet(hoja_nombre)

        # Escribir encabezados
        for i, header in enumerate(encabezados, start=1):
            hoja.cell(row=1, column=i, value=header)

        # Escribir datos
        row_count = 2
        for _, combined_data in data_dict.items():
            for j, header in enumerate(encabezados, start=1):
                hoja.cell(row=row_count, column=j, value=str(combined_data.get(header, "")))
            row_count += 1

        # Guardar cambios
        libro.save(ruta_archivo)

        print(f"[ExcelHandler] Reporte combinado guardado en hoja '{hoja_nombre}'")

    except Exception as e:
        raise IOError(f"Error al guardar reporte combinado en Excel: {e}")
